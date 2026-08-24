#!/usr/bin/env python3
"""PDF → XLSX using PyMuPDF's table detection.

Detects real tables on every page (page.find_tables) and writes each table to
its own worksheet. When no tables are found, falls back to a per-page text
sheet so the result is never empty.

Usage: python3 pdf2excel.py <input.pdf> <output.xlsx>
Prints "TABLES <n>" on success for the route's engine stats.
"""
import sys

import fitz  # PyMuPDF
from openpyxl import Workbook


def main() -> None:
    if len(sys.argv) < 3:
        print("usage: pdf2excel.py <input.pdf> <output.xlsx>", file=sys.stderr)
        sys.exit(2)

    input_path = sys.argv[1]
    output_path = sys.argv[2]

    doc = fitz.open(input_path)
    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet

    table_count = 0

    for page_index in range(len(doc)):
        page = doc[page_index]
        try:
            tables = page.find_tables()
        except Exception:
            tables = []
        for table_index, table in enumerate(tables):
            try:
                rows = table.extract()
            except Exception:
                rows = []
            if not rows:
                continue
            sheet_title = f"P{page_index + 1}_T{table_index + 1}"[:31]
            ws = wb.create_sheet(title=sheet_title)
            for row in rows:
                ws.append([("" if cell is None else str(cell)) for cell in row])
            table_count += 1

    if table_count == 0:
        ws = wb.create_sheet(title="Text")
        for page_index in range(len(doc)):
            ws.append([f"Page {page_index + 1}"])
            text = doc[page_index].get_text().strip()
            for line in text.splitlines():
                ws.append([line])

    wb.save(output_path)
    print(f"TABLES {table_count}")


if __name__ == "__main__":
    main()
